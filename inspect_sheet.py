import urllib.request
import os
import openpyxl

URL = "https://docs.google.com/spreadsheets/d/1Ybhp3gQvH1vA8-xBiteq18VgObN3vggv/export?format=xlsx"
OUTPUT = "ganhos_economicos.xlsx"

def main():
    print("Downloading spreadsheet...")
    try:
        urllib.request.urlretrieve(URL, OUTPUT)
        print(f"Downloaded spreadsheet to {OUTPUT}")
    except Exception as e:
        print(f"Failed to download: {e}")
        return

    if not os.path.exists(OUTPUT):
        print("File doesn't exist.")
        return

    try:
        wb = openpyxl.load_workbook(OUTPUT, read_only=True)
        print("Sheet names:", wb.sheetnames)
        if "Valor Recebido" in wb.sheetnames:
            sheet = wb["Valor Recebido"]
            print("Successfully loaded sheet 'Valor Recebido'")
            # Print first 5 rows to inspect structure
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= 10:
                    break
                print(f"Row {i}: {row}")
        else:
            print("Tab 'Valor Recebido' not found in workbook.")
    except Exception as e:
        print(f"Error reading file: {e}")

if __name__ == "__main__":
    main()
